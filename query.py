"""查询指定学校、专业、省份、年份的录取分数"""
import os, json, sqlite3, requests

# ========== 参数 ==========
SCHOOL_NAME = "郑州大学"
PROVINCE = "山东"
WANT_MAJOR = "土木类"
YEAR = 2025
# =========================

OUTPUT_DIR = "output"

# 1. 查省份 ID
province_id = None
back = requests.get("https://static-data.gaokao.cn/www/2.0/config/81004.json")
all_province = json.loads(back.text)
for key, value in all_province['data'].items():
    if PROVINCE in value.get('provinceName', ''):
        province_id = key
        print(f"省份: {value['provinceName']} (ID: {province_id})")
        break
if not province_id:
    print(f"找不到省份: {PROVINCE}")
    exit(1)

# 2. 查学校 ID
with open("schoolid.json", "r", encoding="utf-8") as f:
    school_data = json.load(f)
school_id = None
school_fullname = None
for s in school_data['data']['school']:
    if SCHOOL_NAME in s['name']:
        school_id = s['school_id']
        school_fullname = s['name']
        print(f"学校: {school_fullname} (ID: {school_id})")
        break
if not school_id:
    print(f"找不到学校: {SCHOOL_NAME}")
    exit(1)

# 3. 先从缓存读
conn = sqlite3.connect('score.db')
conn.execute('''CREATE TABLE IF NOT EXISTS score_data
    (schoolId INTEGER, provinceId INTEGER, year INTEGER, data TEXT,
     PRIMARY KEY (schoolId, provinceId, year))''')
conn.commit()
cur = conn.execute('SELECT data FROM score_data WHERE schoolId=? AND provinceId=? AND year=?',
                   (int(school_id), int(province_id), YEAR))
row = cur.fetchone()
if row:
    all_data = (json.loads(row[0]) if isinstance(row[0], str) else row[0]).get("item", [])
    print(f"从缓存读取: {len(all_data)} 条")
else:
    # 拉取API
    headers = {"User-Agent": "Mozilla/5.0 (compatible; Baiduspider-render/2.0; +http://www.baidu.com/search/spider.html)"}
    all_data = []
    page = 0
    while True:
        page += 1
        params = {"local_province_id": province_id, "page": page, "school_id": school_id,
                  "size": "10", "uri": "apidata/api/gk/score/special", "year": YEAR}
        resp = requests.get("https://api.zjzw.cn/web/api/", params=params, headers=headers)
        data = resp.json()
        if data.get("code") != "0000":
            print(f"API: {data.get('message')}")
            break
        d = data.get("data", {})
        if isinstance(d, str): break
        batch = d.get("item", [])
        if not batch: break
        all_data.extend(batch)
        print(f"已获取第{page}页，累计{len(all_data)}条")
    if all_data:
        conn.execute('INSERT OR REPLACE INTO score_data VALUES (?,?,?,?)',
                     (int(school_id), int(province_id), YEAR,
                      json.dumps({"item": all_data}, ensure_ascii=False)))
        conn.commit()
conn.close()

# 4. 筛选目标专业
matched = [m for m in all_data if WANT_MAJOR in m.get("spname", "")]
display_data = matched if matched else all_data

print(f"\n{'='*60}")
print(f"  {school_fullname} | {PROVINCE} | {YEAR}年")
if matched:
    print(f"  匹配「{WANT_MAJOR}」的专业: {len(matched)} 个")
else:
    print(f"  共 {len(all_data)} 个招生专业（未匹配到「{WANT_MAJOR}」）")
print(f"{'='*60}\n")

# 5. 终端输出（块状格式，不截断）
SEP = "-" * 60
for i, m in enumerate(display_data, 1):
    name = m.get('spname', '')
    score = m.get('min', '-')
    rank = m.get('min_section', '-')
    subject = m.get('sp_info', '-')
    print(f"[{i}] {name}")
    print(f"    最低分: {score}    最低位次: {rank}    选科要求: {subject}")
    print(SEP)

# 6. 保存文本文件
os.makedirs(OUTPUT_DIR, exist_ok=True)
filename = f"{YEAR}_{school_fullname}_{WANT_MAJOR}_{PROVINCE}.txt"
filepath = os.path.join(OUTPUT_DIR, filename)

with open(filepath, "w", encoding="utf-8") as f:
    f.write(f"{school_fullname} | {PROVINCE} | {YEAR}年\n")
    if matched:
        f.write(f"匹配「{WANT_MAJOR}」的专业: {len(matched)} 个\n")
    else:
        f.write(f"共 {len(all_data)} 个招生专业\n")
    f.write("=" * 60 + "\n\n")
    for i, m in enumerate(display_data, 1):
        name = m.get('spname', '')
        score = m.get('min', '-')
        rank = m.get('min_section', '-')
        subject = m.get('sp_info', '-')
        f.write(f"[{i}] {name}\n")
        f.write(f"    最低分: {score}    最低位次: {rank}    选科要求: {subject}\n")
        f.write(SEP + "\n")

# Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"{school_fullname}"

    headers = ["序号", "专业名称", "最低分", "最低位次", "选科要求"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        c.alignment = Alignment(horizontal="center")

    for i, m in enumerate(display_data, 1):
        ws.append([i, m.get('spname', ''), m.get('min', '-'),
                   m.get('min_section', '-'), m.get('sp_info', '-')])

    for col in ws.columns:
        max_w = 0
        for cell in col:
            if cell.value:
                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_w = max(max_w, w)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 60)

    xl_fp = os.path.join(OUTPUT_DIR, f"{YEAR}_{school_fullname}_{WANT_MAJOR}_{PROVINCE}.xlsx")
    wb.save(xl_fp)
    print(f"结果已保存到 {filepath}")
    print(f"Excel 已保存到 {xl_fp}")
except ImportError:
    print(f"结果已保存到 {filepath}")
