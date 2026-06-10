"""山东高考志愿填报：指定院校+专业，冲稳保排序，一校一专业一行"""
import os, re, json, time, sqlite3, requests, unicodedata

# ========== 参数 ==========
PROVINCE = "山东"
MY_RANK = 15000                        # 你的全省排名
WANT_MAJORS = [                        # 目标专业
]
SCHOOLS = ["山东大学", "郑州大学"]       # 目标院校
PRIMARY_YEAR = 2025                   # 冲稳保参考年份
HISTORY_YEARS = [2023, 2024, 2025]     # 展示用历年数据
CHONG_UP = 10000                       # 冲：录取位次比我高0~10000名
WEN_RANGE = 5000                       # 稳：位次差±5000以内
BAO_DOWN = 15000                       # 保：录取位次比我低不超过15000
SHOW_ALL = True                        # True=显示所有匹配（包括超出范围的）
# ==========================

OUTPUT_DIR = "output"
CACHE_DB = "score.db"

# ---- 工具 ----
NOISE = [
    r'[（(]?不招[收]?.*?[）)]', r'[（(]?色盲.*?[）)]', r'[（(]?色弱.*?[）)]',
    r'[（(]?单色识别.*?[）)]', r'[（(]?办学地点[：:].*?[）)]',
    r'[（(]?主校区.*?[）)]', r'[（(]?北校[园区].*?[）)]', r'[（(]?南校[园区].*?[）)]',
    r'[（(]?东校[园区].*?[）)]', r'[（(]?西校[园区].*?[）)]', r'[（(]?新校区.*?[）)]',
]
def strip_n(s):
    for p in NOISE: s = re.sub(p, '', s)
    return re.sub(r'\s+', '', re.sub(r'[（(]{2,}', '（', re.sub(r'[）)]{2,}', '）', s)))

def dw(s):
    w = 0
    for c in s:
        e = unicodedata.east_asian_width(c)
        w += 2 if e in ('F','W','A') else 1
    return w
def pad(s, w):
    n = w - dw(s)
    return s + ' ' * max(n, 0)

# 1. 省份ID
hd = {"User-Agent": "Mozilla/5.0 (compatible; Baiduspider-render/2.0)"}
r = requests.get("https://static-data.gaokao.cn/www/2.0/config/81004.json", headers=hd)
pid = pname = None
for k, v in r.json()['data'].items():
    if PROVINCE in v.get('provinceName', ''): pid, pname = k, v['provinceName']; break
print(f"省份: {pname} ({pid})")

# 2. 学校ID
with open("schoolid.json") as f:
    all_s = json.load(f)['data']['school']
name2id = {s['name']: s['school_id'] for s in all_s}

target = []
for sname in SCHOOLS:
    for k, v in name2id.items():
        if sname in k:
            target.append((k, v))
            break
print(f"匹配到 {len(target)} 所院校")

# 3. 从缓存/API获取数据
conn = sqlite3.connect(CACHE_DB)
conn.execute('''CREATE TABLE IF NOT EXISTS score_data
    (schoolId INTEGER, provinceId INTEGER, year INTEGER, data TEXT,
     PRIMARY KEY (schoolId, provinceId, year))''')
conn.commit()

def get_data(sid, yr):
    """先从缓存取，没有再请求API"""
    sid_i = int(sid)
    cur = conn.execute('SELECT data FROM score_data WHERE schoolId=? AND provinceId=? AND year=?',
                       (sid_i, int(pid), yr))
    row = cur.fetchone()
    if row:
        d = json.loads(row[0]) if isinstance(row[0], str) else row[0]
        return d.get("item", [])

    # API请求
    items = []
    page = 0
    retry = 0
    while True:
        page += 1
        p = {"local_province_id": pid, "page": page, "school_id": sid,
             "size": "10", "uri": "apidata/api/gk/score/special", "year": yr}
        try:
            resp = requests.get("https://api.zjzw.cn/web/api/", params=p, headers=hd, timeout=10)
            data = resp.json()
            code = data.get("code", "")
            if code == "1069":  # 限流
                retry += 1
                if retry <= 3:
                    wait = 30 * retry
                    print(f"\n    限流，等待{wait}秒...", end="\r")
                    time.sleep(wait)
                    page -= 1  # 重试当前页
                    continue
                break
            if code != "0000":
                break
            d = data.get("data", {})
            if isinstance(d, str): break
            batch = d.get("item", [])
            if not batch: break
            items.extend(batch)
            retry = 0
        except: break
    if items:
        conn.execute('INSERT OR REPLACE INTO score_data VALUES (?,?,?,?)',
                     (sid_i, int(pid), yr, json.dumps({"item": items}, ensure_ascii=False)))
        conn.commit()
    return items

# 4. 核心：按 PRIMARY_YEAR 筛选匹配专业
print(f"\n查询 {PRIMARY_YEAR} 年数据...")
results = []
for sname, sid in target:
    print(f"  {sname}", end="\r")
    majors = get_data(sid, PRIMARY_YEAR)
    for m in majors:
        spname = m.get('spname', '')
        if WANT_MAJORS and not any(w in spname for w in WANT_MAJORS):
            continue
        try:
            rk = int(m.get('min_section', 0))
        except: continue
        if rk == 0: continue
        results.append({
            'school': sname, 'major': spname, 'sid': sid,
            'p_rank': rk, 'p_score': m.get('min', '-'),
            'sp_info': m.get('sp_info', '-'), 'history': {}
        })
print(f"\n  匹配: {len(results)} 条")

# 5. 补充历年数据
if results:
    print(f"\n补充 {HISTORY_YEARS} 历年数据...")
    for yr in HISTORY_YEARS:
        if yr == PRIMARY_YEAR:
            for r in results:
                r['history'][yr] = (r['p_score'], str(r['p_rank']))
            continue
        for r in results:
            for m in get_data(r['sid'], yr):
                if strip_n(m.get('spname', '')) == strip_n(r['major']):
                    r['history'][yr] = (m.get('min', '-'), m.get('min_section', '-'))
                    break
            if yr not in r['history']:
                r['history'][yr] = ('-', '-')
        time.sleep(1)

conn.close()

# 6. 冲稳保分类
chong, wen, bao = [], [], []
other = []
for r in results:
    diff = MY_RANK - r['p_rank']
    if 0 < diff <= CHONG_UP:
        r['cat'] = '冲'; chong.append(r)
    elif abs(diff) <= WEN_RANGE:
        r['cat'] = '稳'; wen.append(r)
    elif -BAO_DOWN <= diff < 0:
        r['cat'] = '保'; bao.append(r)
    else:
        if SHOW_ALL:
            r['cat'] = '-'; other.append(r)

for lst in [chong, wen, bao]:
    lst.sort(key=lambda r: r['p_rank'])
all_r = chong + wen + bao + other
print(f"\n冲: {len(chong)}  稳: {len(wen)}  保: {len(bao)}  其他: {len(other)}  共: {len(all_r)}")

# 7. 格式化输出
COL_W, SCH_W, D_W = 34, 16, 16

def section(title, items):
    lines = [f"\n  {title}  ({len(items)} 条)", "-" * 120]
    hdr = pad("专业名称", COL_W) + " | " + pad("院校", SCH_W) + " | "
    for yr in HISTORY_YEARS:
        hdr += pad(f"{yr} 最低分/位次", D_W) + " | "
    hdr += "选科要求"
    lines.append(hdr); lines.append("-" * 120)
    for r in items:
        row = pad(r['major'], COL_W) + " | " + pad(r['school'], SCH_W) + " | "
        for yr in HISTORY_YEARS:
            sc, rk = r['history'].get(yr, ('-', '-'))
            row += pad(f"{sc} / {rk}", D_W) + " | "
        row += r.get('sp_info', '-')
        lines.append(row)
    return lines

out = [
    f"志愿填报参考  |  {pname}  |  排名: {MY_RANK}  |  参考年: {PRIMARY_YEAR}",
    f"目标: {', '.join(WANT_MAJORS) if WANT_MAJORS else '全部专业'}",
    f"冲稳保: 冲 +{CHONG_UP} / 稳 ±{WEN_RANGE} / 保 +{BAO_DOWN}",
    f"共 {len(all_r)} 条  (冲{len(chong)} · 稳{len(wen)} · 保{len(bao)} · 其他{len(other)})",
    "=" * 120,
]
out += section("🔴 冲 — 需要冲一冲", chong)
out += section("🟡 稳 — 比较有把握", wen)
out += section("🟢 保 — 基本能录取", bao)
if other:
    out += section("⚪ 其他 — 超出冲稳保范围，仅供参考", other)

for l in out: print(l)

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 保存 TXT
fp = os.path.join(OUTPUT_DIR, f"志愿填报_{pname}_{MY_RANK}名.txt")
with open(fp, "w", encoding="utf-8") as f:
    for l in out: f.write(l + "\n")

# 保存 Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "志愿填报"

    # 表头
    headers = ["冲稳保", "专业名称", "院校"]
    for yr in HISTORY_YEARS:
        headers += [f"{yr}最低分", f"{yr}最低位次"]
    headers += ["选科要求"]
    ws.append(headers)

    # 样式
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    cat_fills = {
        "冲": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "稳": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "保": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "-": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    for r in all_r:
        row_data = [r['cat'], r['major'], r['school']]
        for yr in HISTORY_YEARS:
            sc, rk = r['history'].get(yr, ('-', '-'))
            row_data += [sc, rk]
        row_data += [r.get('sp_info', '-')]
        ws.append(row_data)

        # 冲稳保列着色
        last_row = ws.max_row
        ws.cell(row=last_row, column=1).fill = cat_fills.get(r['cat'], PatternFill())
        ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center")

    # 自适应列宽
    for col in ws.columns:
        max_w = 0
        for cell in col:
            if cell.value:
                # CJK字符算2宽
                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_w = max(max_w, w)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

    xl_fp = os.path.join(OUTPUT_DIR, f"志愿填报_{pname}_{MY_RANK}名.xlsx")
    wb.save(xl_fp)
    print(f"-> {fp}")
    print(f"-> {xl_fp}")
except ImportError:
    print(f"-> {fp}")
    print("  (安装 openpyxl 可生成 Excel: pip install openpyxl)")
