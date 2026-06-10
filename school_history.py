"""查询指定院校近三年在某个省份的全部专业录取分数"""
import os, re, unicodedata, requests, json

# ========== 参数 ==========
SCHOOL_NAME = "山东大学"
PROVINCE = "山东"
YEARS = [2023, 2024, 2025]
# =========================

OUTPUT_DIR = "output"

# ---- 噪声模式 ----
NOISE = [
    r'[（(]?不招[收]?.*?[）)]', r'[（(]?色盲.*?[）)]', r'[（(]?色弱.*?[）)]',
    r'[（(]?单色识别.*?[）)]', r'[（(]?办学地点[：:].*?[）)]',
    r'[（(]?主校区.*?[）)]', r'[（(]?北校[园区].*?[）)]', r'[（(]?南校[园区].*?[）)]',
    r'[（(]?东校[园区].*?[）)]', r'[（(]?西校[园区].*?[）)]', r'[（(]?新校区.*?[）)]',
]

def core(name):
    for p in NOISE:
        name = re.sub(p, '', name)
    return re.sub(r'[（(]{2,}', '（', re.sub(r'[）)]{2,}', '）', re.sub(r'\s+', '', name.strip())))

# 1-2. ID
resp = requests.get("https://static-data.gaokao.cn/www/2.0/config/81004.json")
data = json.loads(resp.text)['data']
pid = pname = None
for k, v in data.items():
    if PROVINCE in v.get('provinceName', ''): pid, pname = k, v['provinceName']; break
print(f"省份: {pname} ({pid})")

with open("schoolid.json") as f:
    schools = json.load(f)['data']['school']
sid = sname = None
for s in schools:
    if SCHOOL_NAME in s['name']: sid, sname = s['school_id'], s['name']; break
print(f"学校: {sname} ({sid})")

# 3. 拉取（优先缓存）
import sqlite3
conn = sqlite3.connect('score.db')
conn.execute('''CREATE TABLE IF NOT EXISTS score_data
    (schoolId INTEGER, provinceId INTEGER, year INTEGER, data TEXT,
     PRIMARY KEY (schoolId, provinceId, year))''')
conn.commit()

hd = {"User-Agent": "Mozilla/5.0 (compatible; Baiduspider-render/2.0; +http://www.baidu.com/search/spider.html)"}
ydata = {}
for yr in YEARS:
    cur = conn.execute('SELECT data FROM score_data WHERE schoolId=? AND provinceId=? AND year=?',
                       (int(sid), int(pid), yr))
    row = cur.fetchone()
    if row:
        all_d = (json.loads(row[0]) if isinstance(row[0], str) else row[0]).get("item", [])
        print(f"  {yr}年: {len(all_d)} 个专业 (缓存)")
    else:
        all_d, page = [], 0
        while True:
            page += 1
            p = {"local_province_id": pid, "page": page, "school_id": sid,
                 "size": "10", "uri": "apidata/api/gk/score/special", "year": yr}
            r = requests.get("https://api.zjzw.cn/web/api/", params=p, headers=hd)
            data = r.json()
            if data.get("code") != "0000":
                print(f"  {yr}年: {data.get('message')}")
                break
            d = data.get("data", {})
            if isinstance(d, str): break
            items = d.get("item", [])
            if not items: break
            all_d.extend(items)
        if all_d:
            conn.execute('INSERT OR REPLACE INTO score_data VALUES (?,?,?,?)',
                         (int(sid), int(pid), yr, json.dumps({"item": all_d}, ensure_ascii=False)))
            conn.commit()
        print(f"  {yr}年: {len(all_d)} 个专业")
    ydata[yr] = all_d
conn.close()

# 4. 合并
merged = {}
for yr, majors in ydata.items():
    for m in majors:
        ck = core(m.get('spname', ''))
        if ck not in merged:
            merged[ck] = {'names': [], 'years': {}, 'sp': m.get('sp_info', '-')}
        nm = m['spname']
        if nm not in merged[ck]['names']: merged[ck]['names'].append(nm)
        sp = m.get('sp_info', '-')
        if sp != '-' and merged[ck]['sp'] == '-': merged[ck]['sp'] = sp
        merged[ck]['years'][yr] = (m.get('min', '-'), m.get('min_section', '-'))

for ck, v in merged.items():
    uq = list(dict.fromkeys(v['names']))
    v['dname'] = uq[0] if len(uq) == 1 else ' / '.join(uq)

# 5. 排序
def sk(item):
    yd = item[1]['years']
    for y in reversed(YEARS):
        if y in yd:
            try: return int(yd[y][1])
            except: pass
    return 9999999

majors = sorted(merged.items(), key=sk)
print(f"  合并: {len(majors)} 个专业")

# 6. 构建表格 —— 每列按字符数（len）补空格，强行对齐
rows = []
rows.append(["专业名称"] + [f"{y}年" for y in YEARS])
rows.append([""] + ["最低分 / 最低位次"] * len(YEARS))
for ck, v in majors:
    cells = [v['dname']]
    for y in YEARS:
        if y in v['years']:
            sc, rk = v['years'][y]
            cells.append(f"{sc} / {rk}")
        else:
            cells.append("-")
    rows.append(cells)

# 每列最大显示宽度（CJK=2, ASCII=1）
def dw(s):
    w = 0
    for c in s:
        e = unicodedata.east_asian_width(c)
        w += 2 if e in ('F','W','A') else 1
    return w

GAP = 3
cols = len(rows[0])
char_w = [0] * cols
for row in rows:
    for i, cell in enumerate(row):
        char_w[i] = max(char_w[i], dw(cell))
char_w = [w + GAP for w in char_w]

def pad_char(s, w):
    """按显示宽度补空格"""
    need = w - dw(s)
    return s + ' ' * max(need, 0)

# 终端输出
print(f"\n{'='*70}")
print(f"  {sname} | {pname}  ({YEARS[0]}-{YEARS[-1]})")
print(f"  {len(majors)} 个专业")
print(f"{'='*70}\n")

for row in rows:
    line = " | ".join(pad_char(cell, char_w[i]) for i, cell in enumerate(row))
    print(line)

print(f"\n{'='*60}\n选科要求\n{'='*60}")
for ck, v in majors:
    if v.get('sp') and v['sp'] != '-':
        print(f"  {v['dname']}: {v['sp']}")

# 7. 保存
os.makedirs(OUTPUT_DIR, exist_ok=True)
yr_str = f"{YEARS[0]}-{YEARS[-1]}"
fp = os.path.join(OUTPUT_DIR, f"{sname}_{pname}_{yr_str}.txt")

with open(fp, "w", encoding="utf-8") as f:
    f.write(f"{sname} | {pname}  ({yr_str})\n")
    f.write(f"{len(majors)} 个专业\n\n")
    for row in rows:
        f.write(" | ".join(pad_char(cell, char_w[i]) for i, cell in enumerate(row)) + "\n")
    f.write(f"\n{'='*60}\n选科要求\n{'='*60}\n")
    for ck, v in majors:
        if v.get('sp') and v['sp'] != '-':
            f.write(f"  {v['dname']}: {v['sp']}\n")

# Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sname}"

    headers = ["专业名称"]
    for yr in YEARS:
        headers += [f"{yr}最低分", f"{yr}最低位次"]
    headers += ["选科要求"]
    ws.append(headers)

    hdr_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, size=11)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for ck, v in majors:
        row_data = [v['dname']]
        for yr in YEARS:
            if yr in v['years']:
                sc, rk = v['years'][yr]
            else:
                sc, rk = '-', '-'
            row_data += [sc, rk]
        row_data += [v.get('sp', '-')]
        ws.append(row_data)

    for col in ws.columns:
        max_w = 0
        for cell in col:
            if cell.value:
                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_w = max(max_w, w)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 50)

    xl_fp = os.path.join(OUTPUT_DIR, f"{sname}_{pname}_{yr_str}.xlsx")
    wb.save(xl_fp)
    print(f"-> {fp}")
    print(f"-> {xl_fp}")
except ImportError:
    print(f"-> {fp}")
